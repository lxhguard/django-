"""
    mysql_to_excel(username, password, database_name, port)
    参数：username：数据库用户名
          password：数据库用户密码
          database_name：要存入的库的名称
          port：数据库端口
    功能：将数据库相应库的num张以table+n为名称表存入excel中，保存到当前文件夹
    返回值：void

    date:2018.11.24
    author:Benjamin
"""
import pymysql
import openpyxl


def mysql_to_excel(username, password, database_name, port):
    # 创建工作簿,设置编码
    one_excel = openpyxl.Workbook()
    # 连接到数据库
    db = pymysql.connect(host="localhost",
                         user=username,
                         password=password,
                         db=database_name,
                         port=port,
                         autocommit=True)
    # 游标 相当与输入
    cur = db.cursor()
    sql = "show tables like 'as_table%';"
    cur.execute(sql)
    result = cur.fetchall()
    for i in result:
        write_one_table(one_excel, i[0], cur)
        db.commit()
    one_excel.remove(one_excel["Sheet"])
    one_excel.save("50tables.xlsx")
    db.close()


def write_one_table(excel_name, table_name, cur):
    one_sheet = excel_name.create_sheet(title = table_name)
    sql = "select * from %s;" % table_name
    cur.execute(sql)
    result = cur.fetchall()
    row = 1
    for i in result:
        for col in range(1, len(i)+1):
            one_sheet.cell(row=row, column=col).value = i[col-1]
        row += 1


if __name__ == '__main__':
    mysql_to_excel("root", "mysql123", "test", 3306)

